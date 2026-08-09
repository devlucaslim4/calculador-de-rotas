"""Tratamento, métricas, auditoria e exportação da análise de quilometragem."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_ROUTE_KM = 2_000.0

ALIASES = {
    "id": ("ID RESPOSTA", "ID REPOSTA", "ID DA RESPOSTA", "RESPOSTA ID", "ID"),
    "usuario": ("USUARIO", "USUÁRIO", "NOME USUARIO", "NOME DO USUARIO", "COLABORADOR", "MOTORISTA"),
    "unidade": ("UNIDADE", "FILIAL", "BASE", "LOCALIDADE"),
    "regiao": ("REGIAO", "REGIÃO", "REGIONAL"),
    "motivo": ("MOTIVO", "FINALIDADE", "TIPO DE DESLOCAMENTO", "TIPO ROTA"),
    "centro_custo": ("CENTRO DE CUSTO", "CENTRO CUSTO", "CC", "C.C."),
    "status": ("STATUS", "STATUS DA VIAGEM", "SITUACAO", "SITUAÇÃO"),
    "data_inicio": ("DATA INICIO", "DATA DE INICIO", "DATA/HORA INICIO", "INICIO", "DATA DA ROTA"),
    "data_fim": ("DATA FIM", "DATA DE FIM", "DATA CONCLUSAO", "DATA DE CONCLUSAO", "FIM"),
    "odometro_inicial": ("ODOMETRO INICIAL", "HODOMETRO INICIAL", "KM INICIAL", "QUILOMETRAGEM INICIAL"),
    "odometro_final": ("ODOMETRO FINAL", "HODOMETRO FINAL", "KM FINAL", "QUILOMETRAGEM FINAL"),
    "distancia_hodometro_existente": ("DISTANCIA HODOMETRO", "DISTANCIA KM", "KM RODADO", "KM PERCORRIDO"),
    "distancia_gps": ("DISTANCIA GPS",),
    "coord_inicial": ("COORDENADA GPS INICIAL", "COORDENADA GPG INICIAL"),
    "coord_final": ("COORDENADA GPS FINAL",),
}

FILTER_FIELDS = ("usuario", "unidade", "regiao", "motivo", "centro_custo", "status")


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", text).strip().upper()


def resolve_columns(dataframe: pd.DataFrame) -> dict[str, str]:
    normalized = {normalize_header(column): str(column) for column in dataframe.columns}
    resolved: dict[str, str] = {}
    for key, aliases in ALIASES.items():
        for alias in aliases:
            match = normalized.get(normalize_header(alias))
            if match is not None:
                resolved[key] = match
                break
    return resolved


def _number_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")
    cleaned = series.astype("string").str.strip().str.replace(r"[^\d,.-]", "", regex=True)
    comma_decimal = cleaned.str.contains(",", na=False) & ~cleaned.str.contains(r"\.", na=False)
    cleaned.loc[comma_decimal] = cleaned.loc[comma_decimal].str.replace(",", ".", regex=False)
    cleaned.loc[~comma_decimal] = cleaned.loc[~comma_decimal].str.replace(",", "", regex=False)
    return pd.to_numeric(cleaned, errors="coerce")


def _valid_coordinate(value: object) -> bool:
    if pd.isna(value):
        return False
    parts = [part.strip() for part in str(value).split(",")]
    if len(parts) != 2:
        return False
    try:
        latitude, longitude = map(float, parts)
    except (TypeError, ValueError):
        return False
    return -90 <= latitude <= 90 and -180 <= longitude <= 180


@dataclass
class AnalysisData:
    data: pd.DataFrame
    columns: dict[str, str]
    warnings: list[str]

    def unique_routes(self, source: pd.DataFrame | None = None) -> pd.DataFrame:
        frame = self.data if source is None else source
        id_column = self.columns.get("id")
        if id_column:
            valid = frame[id_column].notna() & frame[id_column].astype("string").str.strip().ne("")
            with_id = frame.loc[valid].drop_duplicates(subset=[id_column], keep="first")
            without_id = frame.loc[~valid]
            return pd.concat([with_id, without_id], axis=0).sort_index()
        return frame.copy()


def prepare_analysis(dataframe: pd.DataFrame) -> AnalysisData:
    data = dataframe.copy()
    columns = resolve_columns(data)
    warnings: list[str] = []
    if "id" not in columns:
        warnings.append("Não foi encontrada uma coluna de ID. As linhas serão consideradas rotas individuais.")

    for key in ("data_inicio", "data_fim"):
        column = columns.get(key)
        data[f"__{key}"] = pd.to_datetime(data[column], errors="coerce", dayfirst=True) if column else pd.NaT

    for key in ("odometro_inicial", "odometro_final", "distancia_gps", "distancia_hodometro_existente"):
        column = columns.get(key)
        data[f"__{key}"] = _number_series(data[column]) if column else pd.Series(float("nan"), index=data.index)

    calculated = data["__odometro_final"] - data["__odometro_inicial"]
    valid_odometer = (
        data["__odometro_inicial"].notna()
        & data["__odometro_final"].notna()
        & calculated.ge(0)
        & calculated.le(MAX_ROUTE_KM)
    )
    data["__distancia_hodometro"] = calculated.where(valid_odometer)
    data["__distancia_gps_valida"] = data["__distancia_gps"].where(data["__distancia_gps"].gt(0))
    comparable = data["__distancia_hodometro"].gt(0) & data["__distancia_gps_valida"].gt(0)
    data["__diferenca_percentual"] = (
        (data["__distancia_hodometro"] - data["__distancia_gps_valida"]).abs()
        / data["__distancia_gps_valida"]
        * 100
    ).where(comparable)
    data["__duracao_horas"] = (
        (data["__data_fim"] - data["__data_inicio"]).dt.total_seconds() / 3600
    )
    return AnalysisData(data, columns, warnings)


def apply_filters(analysis: AnalysisData, selections: dict[str, object]) -> pd.DataFrame:
    filtered = analysis.data.copy()
    start, end = selections.get("periodo", (None, None))
    if start is not None:
        filtered = filtered[filtered["__data_inicio"].dt.date >= start]
    if end is not None:
        filtered = filtered[filtered["__data_inicio"].dt.date <= end]
    for key in FILTER_FIELDS:
        values = selections.get(key)
        column = analysis.columns.get(key)
        if column and values:
            filtered = filtered[filtered[column].astype("string").isin(values)]
    return filtered


def calculate_metrics(analysis: AnalysisData, filtered: pd.DataFrame, inconsistency_rows: set[int]) -> dict[str, object]:
    routes = analysis.unique_routes(filtered)
    user = analysis.columns.get("usuario")
    unit = analysis.columns.get("unidade")
    status = analysis.columns.get("status")
    gps = routes["__distancia_gps_valida"]
    status_text = routes[status].astype("string").map(normalize_header) if status else pd.Series("", index=routes.index)
    completed = status_text.str.contains("CONCLUID|FINALIZ|CALCULAD", regex=True, na=False).sum() if status else None
    in_progress = status_text.str.contains("ANDAMENTO|EM CURSO|INICIAD|PENDENTE", regex=True, na=False).sum() if status else None
    total_km = gps.sum(min_count=1)
    return {
        "Total de rotas únicas": len(routes),
        "Total de quilômetros válidos": None if pd.isna(total_km) else float(total_km),
        "Usuários ativos": routes[user].dropna().astype("string").str.strip().replace("", pd.NA).nunique() if user else None,
        "Unidades atendidas": routes[unit].dropna().astype("string").str.strip().replace("", pd.NA).nunique() if unit else None,
        "Média de quilômetros por rota": None if gps.dropna().empty else float(gps.mean()),
        "Rotas concluídas": int(completed) if completed is not None else None,
        "Rotas em andamento": int(in_progress) if in_progress is not None else None,
        "Registros com inconsistências": len(set(filtered.index) & inconsistency_rows),
    }


def build_audit(analysis: AnalysisData, divergence_limit: float = 20, duration_limit: float = 8) -> pd.DataFrame:
    data, columns = analysis.data, analysis.columns
    records: list[dict[str, object]] = []
    id_col, user_col, unit_col = columns.get("id"), columns.get("usuario"), columns.get("unidade")

    def add(index: int, kind: str, value: object, note: str) -> None:
        row = data.loc[index]
        records.append({
            "Índice da linha": index,
            "ID da resposta": row.get(id_col) if id_col else None,
            "Usuário": row.get(user_col) if user_col else None,
            "Unidade": row.get(unit_col) if unit_col else None,
            "Data": row.get("__data_inicio"),
            "Tipo da inconsistência": kind,
            "Valor encontrado": value,
            "Observação": note,
        })

    if id_col:
        duplicated = data[id_col].notna() & data.duplicated(id_col, keep=False)
        for index in data.index[duplicated]:
            add(index, "Rota duplicada pelo ID", data.at[index, id_col], "O ID aparece em mais de uma linha; centros de custo foram preservados.")

    initial_odo, final_odo = columns.get("odometro_inicial"), columns.get("odometro_final")
    for index in data.index:
        if initial_odo and pd.isna(data.at[index, "__odometro_inicial"]):
            add(index, "Hodômetro inicial ausente", data.at[index, initial_odo], "Não foi possível calcular a distância do hodômetro.")
        if final_odo and pd.isna(data.at[index, "__odometro_final"]):
            add(index, "Hodômetro final ausente", data.at[index, final_odo], "Não foi possível calcular a distância do hodômetro.")
        if data.at[index, "__odometro_final"] < data.at[index, "__odometro_inicial"]:
            add(index, "Hodômetro final menor que o inicial", data.at[index, "__distancia_hodometro"], "A diferença negativa é inválida.")
        raw_difference = data.at[index, "__odometro_final"] - data.at[index, "__odometro_inicial"]
        if pd.notna(raw_difference) and raw_difference > MAX_ROUTE_KM:
            add(index, "Distância fora do padrão", raw_difference, f"O limite adotado é {MAX_ROUTE_KM:,.0f} km por rota.")
        if data.at[index, "__distancia_hodometro"] == 0:
            add(index, "Distância igual a zero", 0, "Os hodômetros inicial e final são iguais.")

        for key, label in (("coord_inicial", "Coordenada inicial ausente ou inválida"), ("coord_final", "Coordenada final ausente ou inválida")):
            column = columns.get(key)
            if column and not _valid_coordinate(data.at[index, column]):
                add(index, label, data.at[index, column], "Use o formato latitude, longitude dentro dos limites válidos.")
        if pd.isna(data.at[index, "__distancia_gps_valida"]):
            add(index, "Distância GPS ausente", data.at[index, columns["distancia_gps"]] if "distancia_gps" in columns else None, "Não há distância rodoviária válida para esta linha.")
        difference = data.at[index, "__diferenca_percentual"]
        if pd.notna(difference) and difference > divergence_limit:
            add(index, "Divergência elevada entre GPS e hodômetro", round(difference, 2), f"A diferença supera o limite de {divergence_limit:.0f}%.")
        duration = data.at[index, "__duracao_horas"]
        if pd.notna(duration) and duration < 0:
            add(index, "Conclusão anterior ao início", round(duration, 2), "A data de conclusão é anterior à data de início.")
        elif pd.notna(duration) and duration > duration_limit:
            add(index, "Duração excessiva", round(duration, 2), f"A duração supera o limite de {duration_limit:.1f} horas.")

        status_col = columns.get("status")
        status = normalize_header(data.at[index, status_col]) if status_col else ""
        if status and any(term in status for term in ("ANDAMENTO", "EM CURSO", "INICIAD", "PENDENTE")):
            add(index, "Registro em andamento", data.at[index, status_col], "A rota ainda não está concluída.")
        cost_col = columns.get("centro_custo")
        if cost_col:
            value = data.at[index, cost_col]
            if pd.isna(value) or normalize_header(value) in {"", "N/A", "NA", "0", "-"}:
                add(index, "Centro de custo ausente ou inválido", value, "Revise o centro de custo informado.")

        existing = data.at[index, "__distancia_hodometro_existente"]
        recalculated = data.at[index, "__distancia_hodometro"]
        if pd.notna(existing) and pd.notna(recalculated) and abs(existing - recalculated) > 0.01:
            add(index, "Distância de hodômetro divergente", existing, f"Valor recalculado: {recalculated:.2f} km.")

    columns_out = ["Índice da linha", "ID da resposta", "Usuário", "Unidade", "Data", "Tipo da inconsistência", "Valor encontrado", "Observação"]
    return pd.DataFrame(records, columns=columns_out)


def analysis_workbook(
    metrics: dict[str, object],
    filtered: pd.DataFrame,
    audit: pd.DataFrame,
    analysis: AnalysisData,
) -> bytes:
    routes = analysis.unique_routes(filtered)
    user, unit = analysis.columns.get("usuario"), analysis.columns.get("unidade")
    user_summary = _summary(routes, user)
    unit_summary = _summary(routes, unit)
    public_columns = [column for column in filtered.columns if not str(column).startswith("__")]
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame({"Indicador": metrics.keys(), "Valor": metrics.values()}).to_excel(writer, sheet_name="Resumo", index=False)
        user_summary.to_excel(writer, sheet_name="Rotas por usuário", index=False)
        unit_summary.to_excel(writer, sheet_name="Rotas por unidade", index=False)
        filtered[public_columns].to_excel(writer, sheet_name="Dados filtrados", index=False)
        audit.drop(columns=["Índice da linha"], errors="ignore").to_excel(writer, sheet_name="Inconsistências", index=False)
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2563EB")
            for column_cells in worksheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    return output.getvalue()


def _summary(routes: pd.DataFrame, column: str | None) -> pd.DataFrame:
    if not column:
        return pd.DataFrame(columns=["Categoria", "Rotas", "Quilômetros GPS"])
    result = routes.groupby(column, dropna=False).agg(
        Rotas=(column, "size"),
        **{"Quilômetros GPS": ("__distancia_gps_valida", "sum")},
    ).reset_index().rename(columns={column: "Categoria"})
    return result.sort_values("Rotas", ascending=False)
