from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader

from analysis_engine import (
    analysis_workbook,
    apply_filters,
    build_audit,
    calculate_metrics,
    prepare_analysis,
    resolve_columns,
)
from report_export import analysis_pdf


def sample_dataframe():
    return pd.DataFrame({
        " ID REPOSTA ": ["R1", "R1", "R2", "R3"],
        "Usuário": ["Ana", "Ana", "Bruno", "Ana"],
        "Unidade": ["SP", "SP", "RJ", "SP"],
        "Região": ["Sudeste"] * 4,
        "Motivo": ["Visita", "Visita", "Entrega", "Visita"],
        "Centro de Custo": ["100", "200", None, "300"],
        "Status": ["Concluída", "Concluída", "Em andamento", "Concluída"],
        "Data de início": ["01/08/2026 08:00", "01/08/2026 08:00", "02/08/2026 10:00", "03/08/2026 09:00"],
        "Data de conclusão": ["01/08/2026 10:00", "01/08/2026 10:00", None, "03/08/2026 20:00"],
        "Hodômetro inicial": [1000, 1000, 2000, 3000],
        "Hodômetro final": [1100, 1100, None, 2990],
        "DISTÂNCIA GPS": [90, 90, 50, None],
        "COORDENADA GPS INICIAL": ["-23.5, -46.6", "-23.5, -46.6", "inválida", "-23.5, -46.6"],
        "COORDENADA GPS FINAL": ["-22.9, -43.2", "-22.9, -43.2", "-22.9, -43.2", ""],
    })


def test_resolves_known_header_variation_and_deduplicates_routes():
    analysis = prepare_analysis(sample_dataframe())
    assert analysis.columns["id"] == " ID REPOSTA "
    assert len(analysis.unique_routes()) == 3
    assert len(analysis.data) == 4


def test_recalculates_odometer_and_metrics_without_counting_duplicate_route():
    analysis = prepare_analysis(sample_dataframe())
    assert analysis.data.loc[0, "__distancia_hodometro"] == 100
    assert pd.isna(analysis.data.loc[2, "__distancia_hodometro"])
    assert pd.isna(analysis.data.loc[3, "__distancia_hodometro"])
    audit = build_audit(analysis)
    metrics = calculate_metrics(analysis, analysis.data, set(audit["Índice da linha"]))
    assert metrics["Total de rotas únicas"] == 3
    assert metrics["Total de quilômetros válidos"] == 140
    assert metrics["Usuários ativos"] == 2


def test_filters_preserve_cost_center_rows_and_audit_expected_problems():
    analysis = prepare_analysis(sample_dataframe())
    filtered = apply_filters(analysis, {"usuario": ["Ana"], "status": ["Concluída"], "unidade": ["SP"]})
    assert len(filtered) == 3
    audit = build_audit(analysis, divergence_limit=5, duration_limit=8)
    kinds = set(audit["Tipo da inconsistência"])
    assert "Rota duplicada pelo ID" in kinds
    assert "Hodômetro final ausente" in kinds
    assert "Hodômetro final menor que o inicial" in kinds
    assert "Coordenada inicial ausente ou inválida" in kinds
    assert "Duração excessiva" in kinds


def test_missing_id_uses_rows_and_emits_warning():
    frame = sample_dataframe().drop(columns=[" ID REPOSTA "])
    analysis = prepare_analysis(frame)
    assert len(analysis.unique_routes()) == len(frame)
    assert analysis.warnings


def test_analysis_workbook_contains_expected_sheets_and_opens():
    analysis = prepare_analysis(sample_dataframe())
    audit = build_audit(analysis)
    metrics = calculate_metrics(analysis, analysis.data, set(audit["Índice da linha"]))
    result = analysis_workbook(metrics, analysis.data, audit, analysis)
    workbook = load_workbook(BytesIO(result), read_only=True)
    assert workbook.sheetnames == ["Resumo", "Rotas por usuário", "Rotas por unidade", "Dados filtrados", "Inconsistências"]


def test_analysis_pdf_contains_summary_and_audit_pages():
    analysis = prepare_analysis(sample_dataframe())
    audit = build_audit(analysis)
    metrics = calculate_metrics(analysis, analysis.data, set(audit["Índice da linha"]))
    result = analysis_pdf(metrics, analysis.data, audit, analysis, "rotas_teste.xlsx")
    reader = PdfReader(BytesIO(result))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert len(reader.pages) >= 2
    assert "Relatorio consolidado de rotas" in text
    assert "Auditoria dos dados" in text
