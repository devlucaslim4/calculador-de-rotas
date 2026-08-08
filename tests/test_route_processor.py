from io import BytesIO

from openpyxl import Workbook, load_workbook

import pytest

from route_processor import SpreadsheetError, RouteResult, output_filename, parse_coordinate, process_workbook


def make_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rotas"
    sheet.append(["Cliente", " COORDENADA gpg INICIAL ", "coordenada gps final"])
    sheet.append(["Válida", "-23.5505, -46.6333", "-22.9068, -43.1729"])
    sheet.append(["Origem inválida", "91, 0", "-22.9, -43.1"])
    sheet.append(["Destino vazio", "-23.5, -46.6", None])
    extra = workbook.create_sheet("Observações")
    extra["A1"] = "Conteúdo preservado"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def fake_route(origin, destination):
    parsed_origin = parse_coordinate(origin)
    parsed_destination = parse_coordinate(destination)
    if parsed_origin is None:
        return RouteResult(None, "Coordenada inicial inválida", None)
    if parsed_destination is None:
        return RouteResult(None, "Coordenada final inválida", None)
    return RouteResult(432.1, "Calculada", "https://www.google.com/maps/dir/?api=1")


def test_parse_coordinate_validates_format_and_limits():
    assert parse_coordinate(" -23.5, -46.6 ") == (-23.5, -46.6)
    assert parse_coordinate("91, 0") is None
    assert parse_coordinate("0, 181") is None
    assert parse_coordinate("") is None


def test_process_workbook_preserves_sheets_and_writes_results():
    progress = []
    summary = process_workbook(make_workbook(), lambda done, total: progress.append((done, total)), fake_route)
    assert (summary.total, summary.calculated, summary.failed) == (3, 1, 2)
    assert progress[-1] == (3, 3)

    workbook = load_workbook(BytesIO(summary.workbook))
    assert workbook.sheetnames == ["Rotas", "Observações"]
    assert workbook["Observações"]["A1"].value == "Conteúdo preservado"
    sheet = workbook["Rotas"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(2, headers["DISTÂNCIA GPS"]).value == 432.1
    assert sheet.cell(2, headers["STATUS DA ROTA"]).value == "Calculada"
    assert sheet.cell(2, headers["LINK DA ROTA"]).value == "ABRIR ROTA"
    assert sheet.cell(2, headers["LINK DA ROTA"]).hyperlink.target.startswith("https://")
    assert sheet.cell(3, headers["LINK DA ROTA"]).value is None


def test_output_filename_uses_uploaded_name():
    assert output_filename("rotas_junho.xlsx") == "rotas_junho_com_distancia_gps.xlsx"


def test_rejects_corrupted_workbook():
    with pytest.raises(SpreadsheetError, match="corrompido"):
        process_workbook("não é um arquivo xlsx".encode(), route_function=fake_route)


def test_rejects_missing_required_columns():
    workbook = Workbook()
    workbook.active.append(["Cliente", "Endereço"])
    workbook.active.append(["Teste", "Rua A"])
    output = BytesIO()
    workbook.save(output)
    with pytest.raises(SpreadsheetError, match="Coluna"):
        process_workbook(output.getvalue(), route_function=fake_route)
