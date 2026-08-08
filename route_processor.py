"""Processamento de planilhas e consulta de rotas via OSRM."""

from __future__ import annotations

import logging
import math
import re
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Callable
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

LOGGER = logging.getLogger(__name__)
OSRM_URL = "https://router.project-osrm.org/route/v1/driving"
MAX_WORKERS = 4
REQUEST_TIMEOUT = (5, 20)
MAX_RETRIES = 3

INITIAL_HEADER = "COORDENADA GPS INICIAL"
LEGACY_INITIAL_HEADER = "COORDENADA GPG INICIAL"
FINAL_HEADER = "COORDENADA GPS FINAL"
RESULT_HEADERS = ("DISTÂNCIA GPS", "STATUS DA ROTA", "LINK DA ROTA")


class SpreadsheetError(ValueError):
    """Erro amigável relacionado à planilha enviada."""


@dataclass(frozen=True)
class RouteResult:
    distance_km: float | None
    status: str
    link: str | None


@dataclass(frozen=True)
class ProcessingSummary:
    workbook: bytes
    total: int
    calculated: int
    failed: int


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", text).strip().upper()


def parse_coordinate(value: object) -> tuple[float, float] | None:
    if value is None or isinstance(value, bool):
        return None
    text = str(value).strip()
    parts = [part.strip() for part in text.split(",")]
    if len(parts) != 2:
        return None
    try:
        latitude, longitude = map(float, parts)
    except (TypeError, ValueError):
        return None
    if not (math.isfinite(latitude) and math.isfinite(longitude)):
        return None
    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
        return None
    return latitude, longitude


def google_maps_link(origin: tuple[float, float], destination: tuple[float, float]) -> str:
    params = urlencode({
        "api": 1,
        "origin": f"{origin[0]},{origin[1]}",
        "destination": f"{destination[0]},{destination[1]}",
        "travelmode": "driving",
    })
    return f"https://www.google.com/maps/dir/?{params}"


def calculate_route(
    origin_value: object,
    destination_value: object,
    session: requests.Session | None = None,
) -> RouteResult:
    origin = parse_coordinate(origin_value)
    if origin is None:
        return RouteResult(None, "Coordenada inicial inválida", None)
    destination = parse_coordinate(destination_value)
    if destination is None:
        return RouteResult(None, "Coordenada final inválida", None)

    client = session or requests.Session()
    coordinates = f"{origin[1]},{origin[0]};{destination[1]},{destination[0]}"
    url = f"{OSRM_URL}/{coordinates}"
    for attempt in range(MAX_RETRIES):
        try:
            response = client.get(
                url,
                params={"overview": "false", "steps": "false"},
                timeout=REQUEST_TIMEOUT,
                headers={"User-Agent": "Calculador-de-Rotas/1.0"},
            )
            if response.status_code in {429, 500, 502, 503, 504}:
                raise requests.RequestException(f"OSRM temporariamente indisponível ({response.status_code})")
            response.raise_for_status()
            data = response.json()
            if data.get("code") == "NoRoute" or not data.get("routes"):
                return RouteResult(None, "Rota não encontrada", None)
            if data.get("code") != "Ok":
                return RouteResult(None, "Falha na consulta", None)
            distance = round(float(data["routes"][0]["distance"]) / 1000, 2)
            return RouteResult(distance, "Calculada", google_maps_link(origin, destination))
        except (requests.RequestException, ValueError, KeyError, TypeError) as exc:
            LOGGER.warning("Falha na consulta OSRM (tentativa %s/%s): %s", attempt + 1, MAX_RETRIES, exc)
            if attempt < MAX_RETRIES - 1:
                time.sleep(0.75 * (2**attempt))
    return RouteResult(None, "Falha na consulta", None)


def _find_headers(worksheet) -> tuple[int, int, dict[str, int]]:
    headers: dict[str, int] = {}
    for cell in worksheet[1]:
        normalized = normalize_header(cell.value)
        if normalized:
            headers[normalized] = cell.column

    initial = headers.get(normalize_header(INITIAL_HEADER))
    if initial is None:
        initial = headers.get(normalize_header(LEGACY_INITIAL_HEADER))
    final = headers.get(normalize_header(FINAL_HEADER))
    missing = []
    if initial is None:
        missing.append(INITIAL_HEADER)
    if final is None:
        missing.append(FINAL_HEADER)
    if missing:
        raise SpreadsheetError("Coluna(s) obrigatória(s) não encontrada(s): " + ", ".join(missing) + ".")
    return initial, final, headers


def process_workbook(
    file_bytes: bytes,
    progress_callback: Callable[[int, int], None] | None = None,
    route_function: Callable[[object, object], RouteResult] = calculate_route,
) -> ProcessingSummary:
    try:
        workbook = load_workbook(BytesIO(file_bytes))
    except Exception as exc:
        LOGGER.exception("Não foi possível abrir a planilha")
        raise SpreadsheetError("O arquivo não é uma planilha Excel válida ou está corrompido.") from exc

    worksheet = workbook.active
    if worksheet.max_row < 2:
        raise SpreadsheetError("A planilha está vazia. Inclua ao menos uma linha de dados.")
    initial_col, final_col, headers = _find_headers(worksheet)
    total = worksheet.max_row - 1

    result_columns: dict[str, int] = {}
    next_column = worksheet.max_column + 1
    for header in RESULT_HEADERS:
        normalized = normalize_header(header)
        column = headers.get(normalized)
        if column is None:
            column = next_column
            next_column += 1
            worksheet.cell(1, column, header)
        result_columns[header] = column

    rows = [
        (row, worksheet.cell(row, initial_col).value, worksheet.cell(row, final_col).value)
        for row in range(2, worksheet.max_row + 1)
    ]
    results: dict[int, RouteResult] = {}
    completed = 0
    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, max(1, total))) as executor:
        futures = {executor.submit(route_function, origin, destination): row for row, origin, destination in rows}
        for future in as_completed(futures):
            row = futures[future]
            try:
                results[row] = future.result()
            except Exception:
                LOGGER.exception("Erro inesperado ao processar a linha %s", row)
                results[row] = RouteResult(None, "Falha na consulta", None)
            completed += 1
            if progress_callback:
                progress_callback(completed, total)

    calculated = 0
    for row, _, _ in rows:
        result = results[row]
        distance_cell = worksheet.cell(row, result_columns["DISTÂNCIA GPS"])
        status_cell = worksheet.cell(row, result_columns["STATUS DA ROTA"])
        link_cell = worksheet.cell(row, result_columns["LINK DA ROTA"])
        distance_cell.value = result.distance_km
        distance_cell.number_format = "0.00"
        status_cell.value = result.status
        link_cell.value = "ABRIR ROTA" if result.link else None
        link_cell.hyperlink = result.link
        if result.link:
            link_cell.style = "Hyperlink"
        if result.status == "Calculada":
            calculated += 1

    for header, width in (("DISTÂNCIA GPS", 18), ("STATUS DA ROTA", 32), ("LINK DA ROTA", 18)):
        column = result_columns[header]
        header_cell = worksheet.cell(1, column)
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill("solid", fgColor="2563EB")
        worksheet.column_dimensions[get_column_letter(column)].width = width

    output = BytesIO()
    workbook.save(output)
    return ProcessingSummary(output.getvalue(), total, calculated, total - calculated)


def output_filename(original_name: str) -> str:
    safe_name = Path(original_name).name
    stem = Path(safe_name).stem or "planilha"
    return f"{stem}_com_distancia_gps.xlsx"
